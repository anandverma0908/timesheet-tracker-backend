"""
Report Generator — produces TWO Excel files matching your exact existing formats:

1. Monthly Finance Sheet  → matches Oct_25_Engg_time_sheet.xlsx
   Sheets: Raw Data (OCT 25 style), Summary (POD wise), Breakdown (Sheet6 style)

2. Annual FY Engineering Sheet → matches Engineering_Timesheet_FY_2024-2025.xlsx
   Sheets: One per POD — Resource Name | Emp Code | Product | Task | Client | Apr | May | ...
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime
import calendar
import os


# ── Style helpers ────────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

def _font(bold=False, color="000000", size=10, name="Calibri"):
    return Font(bold=bold, color=color, size=size, name=name)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border_thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def _header_cell(ws, row, col, value,
                 bg="1F4E79", fg="FFFFFF", bold=True, size=10,
                 h_align="center", wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = _fill(bg)
    c.font      = _font(bold=bold, color=fg, size=size)
    c.alignment = _align(h_align, "center", wrap)
    c.border    = _border_thin()
    return c

def _data_cell(ws, row, col, value, bg="FFFFFF", bold=False,
               color="000000", h_align="left", num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = _fill(bg)
    c.font      = _font(bold=bold, color=color, size=9)
    c.alignment = _align(h_align, "center")
    c.border    = _border_thin()
    if num_fmt:
        c.number_format = num_fmt
    return c


# ════════════════════════════════════════════════════════════════════════════
# FILE 1 — Monthly Finance Sheet
# Matches: Oct_25_Engg_time_sheet.xlsx
# ════════════════════════════════════════════════════════════════════════════

def generate_monthly_finance_report(rows, tickets, month_label, date_from=None, date_to=None):
    """
    rows    = flat worklog rows (one per log entry, with name/pod/date/feature/type/client/hours/jira)
    tickets = full ticket list (for summary aggregations)
    month_label = e.g. "OCT 25"
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _sheet_raw_data(wb, rows, month_label)
    _sheet_pod_summary(wb, rows, month_label)
    _sheet_breakdown(wb, rows)

    os.makedirs("/tmp/reports", exist_ok=True)
    path = f"/tmp/reports/monthly_timesheet_{month_label.replace(' ', '_')}.xlsx"
    wb.save(path)
    return path


# ── Sheet 1: Raw Data (matches "OCT 25" sheet) ───────────────────────────────

def _sheet_raw_data(wb, rows, month_label):
    ws = wb.create_sheet(month_label)
    ws.freeze_panes = "A2"

    # Header row — matches exactly: Name | POD | Date | Module | Feature | Type | Client | Time spent in (Hours) | JIRA# | Remark
    headers = [
        "Name", "POD", "Date", "Module", "Feature",
        "Type\n(Bug/Feature/Meeting)", "Client",
        "Time spent in (Hours)", "JIRA#", "Remark"
    ]
    widths = [24, 12, 12, 14, 40, 20, 16, 22, 12, 30]

    ws.row_dimensions[1].height = 30
    for col_idx, (h, w) in enumerate(zip(headers, widths), 1):
        _header_cell(ws, 1, col_idx, h, bg="1F4E79", fg="FFFFFF",
                     bold=True, size=10, h_align="center", wrap=True)
        _set_col_width(ws, get_column_letter(col_idx), w)

    # Alternating row colors — light blue / white (matches original style)
    ROW_COLORS = ["EBF3FB", "FFFFFF"]
    for i, row in enumerate(rows, start=2):
        bg = ROW_COLORS[i % 2]
        ws.row_dimensions[i].height = 16

        date_val = row.get("date", "")
        if isinstance(date_val, str) and date_val:
            try:
                date_val = datetime.strptime(date_val, "%Y-%m-%d")
            except Exception:
                pass

        values = [
            row.get("name", ""),
            row.get("pod", ""),
            date_val,
            row.get("module", ""),
            row.get("feature", ""),
            row.get("type", ""),
            row.get("client", ""),
            row.get("hours", 0),
            row.get("jira", ""),
            row.get("remark", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = _data_cell(ws, i, col_idx, val, bg=bg)
            if col_idx == 3 and isinstance(val, datetime):
                cell.number_format = "DD-MMM-YY"
            if col_idx == 8:
                cell.alignment = _align("center", "center")
                cell.font = _font(bold=True, color="1F4E79", size=9)
            if col_idx == 9:
                cell.font = _font(bold=True, color="2E75B6", size=9)

    # Total row
    total_row = len(rows) + 2
    ws.row_dimensions[total_row].height = 18
    tc = ws.cell(row=total_row, column=7, value="Grand Total")
    tc.font = _font(bold=True, color="FFFFFF", size=10)
    tc.fill = _fill("1F4E79")
    tc.alignment = _align("right", "center")
    tc.border = _border_thin()

    tv = ws.cell(row=total_row, column=8, value=f"=SUM(H2:H{total_row-1})")
    tv.font = _font(bold=True, color="FFFFFF", size=10)
    tv.fill = _fill("1F4E79")
    tv.alignment = _align("center", "center")
    tv.border = _border_thin()
    tv.number_format = "0.00"


# ── Sheet 2: POD Summary (matches "Time sheet summarry OCT25") ────────────────

def _sheet_pod_summary(wb, rows, month_label):
    ws = wb.create_sheet(f"Summary {month_label}")

    # Title
    ws.merge_cells("B1:D1")
    title = ws.cell(row=1, column=2, value=f"Timesheet Summary — {month_label}")
    title.font      = _font(bold=True, color="FFFFFF", size=13)
    title.fill      = _fill("1F4E79")
    title.alignment = _align("center", "center")
    ws.row_dimensions[1].height = 28

    # Header — matches original: (blank) | POD | Time spent in (Hours) | Story Points
    ws.row_dimensions[2].height = 20
    for col, (val, w) in enumerate(
        [("", 4), ("POD", 20), (" Time spent in (Hours)", 24), ("  Story Points", 16)],
        start=1
    ):
        _header_cell(ws, 2, col, val, bg="2E75B6", fg="FFFFFF",
                     bold=True, size=10, h_align="center")
        _set_col_width(ws, get_column_letter(col), w)

    # Group by POD
    pod_hours = defaultdict(float)
    for row in rows:
        pod_hours[row.get("pod", "Not Set")] += row.get("hours", 0)

    ROW_COLORS = ["EBF3FB", "FFFFFF"]
    for i, (pod, hours) in enumerate(sorted(pod_hours.items()), start=3):
        bg = ROW_COLORS[i % 2]
        ws.row_dimensions[i].height = 16
        _data_cell(ws, i, 1, None, bg=bg)
        _data_cell(ws, i, 2, pod, bg=bg, bold=True, color="1F4E79")
        _data_cell(ws, i, 3, round(hours, 2), bg=bg, h_align="center",
                   num_fmt="0.00")
        _data_cell(ws, i, 4, None, bg=bg)

    # Grand total row
    total_row = len(pod_hours) + 3
    ws.row_dimensions[total_row].height = 18
    for col in range(1, 5):
        c = ws.cell(row=total_row, column=col)
        c.fill   = _fill("1F4E79")
        c.border = _border_thin()
        c.font   = _font(bold=True, color="FFFFFF", size=10)
    ws.cell(row=total_row, column=2, value="Grand Total").alignment = _align("center", "center")
    tv = ws.cell(row=total_row, column=3, value=f"=SUM(C3:C{total_row-1})")
    tv.number_format = "0.00"
    tv.alignment     = _align("center", "center")
    tv.font          = _font(bold=True, color="FFFFFF", size=10)
    tv.fill          = _fill("1F4E79")
    tv.border        = _border_thin()


# ── Sheet 3: Breakdown (matches "Sheet6" — POD wise + Client wise side by side) ──

def _sheet_breakdown(wb, rows):
    ws = wb.create_sheet("Breakdown")

    # ── POD wise section ──
    pod_type = defaultdict(lambda: defaultdict(float))
    all_types = set()
    for row in rows:
        t = row.get("type", "Feature")
        pod_type[row.get("pod", "Not Set")][t] += row.get("hours", 0)
        all_types.add(t)

    type_cols = sorted(all_types) + ["Grand Total"]

    # Title
    title_end = 1 + len(type_cols)
    ws.merge_cells(f"A1:{get_column_letter(title_end)}1")
    tc = ws.cell(row=1, column=1, value="POD wise")
    tc.font = _font(bold=True, color="FFFFFF", size=11)
    tc.fill = _fill("1F4E79")
    tc.alignment = _align("center", "center")
    ws.row_dimensions[1].height = 22

    # Header row
    ws.row_dimensions[2].height = 18
    _header_cell(ws, 2, 1, "POD", bg="2E75B6", fg="FFFFFF")
    _set_col_width(ws, "A", 20)
    for ci, t in enumerate(type_cols, start=2):
        _header_cell(ws, 2, ci, t, bg="2E75B6", fg="FFFFFF", h_align="center")
        _set_col_width(ws, get_column_letter(ci), 14)

    # Data rows
    pods_sorted = sorted(pod_type.keys())
    ROW_COLORS = ["EBF3FB", "FFFFFF"]
    for ri, pod in enumerate(pods_sorted, start=3):
        bg = ROW_COLORS[ri % 2]
        ws.row_dimensions[ri].height = 16
        _data_cell(ws, ri, 1, pod, bg=bg, bold=True, color="1F4E79")
        row_total = 0
        for ci, t in enumerate(type_cols[:-1], start=2):
            val = round(pod_type[pod].get(t, 0) or 0, 2)
            row_total += val
            _data_cell(ws, ri, ci, val if val else None, bg=bg, h_align="center",
                       num_fmt="0.00")
        _data_cell(ws, ri, len(type_cols) + 1, round(row_total, 2),
                   bg=bg, bold=True, h_align="center", num_fmt="0.00")

    # Grand total row
    gt_row = len(pods_sorted) + 3
    ws.row_dimensions[gt_row].height = 18
    for ci in range(1, len(type_cols) + 2):
        c = ws.cell(row=gt_row, column=ci)
        c.fill = _fill("1F4E79"); c.border = _border_thin()
        c.font = _font(bold=True, color="FFFFFF", size=10)
        c.alignment = _align("center", "center")
    ws.cell(row=gt_row, column=1, value="Grand Total")
    for ci in range(2, len(type_cols) + 2):
        col_letter = get_column_letter(ci)
        ws.cell(row=gt_row, column=ci,
                value=f"=SUM({col_letter}3:{col_letter}{gt_row-1})")

    # ── Client wise section — starts 3 columns to the right ──
    start_col = len(type_cols) + 4

    client_type = defaultdict(lambda: defaultdict(float))
    for row in rows:
        t = row.get("type", "Feature")
        client_type[row.get("client", "Not Set")][t] += row.get("hours", 0)

    # Title
    cl_end = start_col + len(type_cols)
    ws.merge_cells(f"{get_column_letter(start_col)}1:{get_column_letter(cl_end)}1")
    ct = ws.cell(row=1, column=start_col, value="Client  wise")
    ct.font = _font(bold=True, color="FFFFFF", size=11)
    ct.fill = _fill("1F4E79")
    ct.alignment = _align("center", "center")

    # Header row
    _header_cell(ws, 2, start_col, "Client", bg="2E75B6", fg="FFFFFF")
    _set_col_width(ws, get_column_letter(start_col), 22)
    for ci, t in enumerate(type_cols, start=start_col + 1):
        _header_cell(ws, 2, ci, t, bg="2E75B6", fg="FFFFFF", h_align="center")
        _set_col_width(ws, get_column_letter(ci), 14)

    clients_sorted = sorted(client_type.keys())
    for ri, client in enumerate(clients_sorted, start=3):
        bg = ROW_COLORS[ri % 2]
        ws.row_dimensions[ri].height = 16
        _data_cell(ws, ri, start_col, client, bg=bg, bold=True, color="1F4E79")
        row_total = 0
        for ci, t in enumerate(type_cols[:-1], start=start_col + 1):
            val = round(client_type[client].get(t, 0) or 0, 2)
            row_total += val
            _data_cell(ws, ri, ci, val if val else None, bg=bg,
                       h_align="center", num_fmt="0.00")
        _data_cell(ws, ri, start_col + len(type_cols), round(row_total, 2),
                   bg=bg, bold=True, h_align="center", num_fmt="0.00")

    # Grand total
    gt_row2 = len(clients_sorted) + 3
    ws.row_dimensions[gt_row2].height = 18
    for ci in range(start_col, start_col + len(type_cols) + 1):
        c = ws.cell(row=gt_row2, column=ci)
        c.fill = _fill("1F4E79"); c.border = _border_thin()
        c.font = _font(bold=True, color="FFFFFF", size=10)
        c.alignment = _align("center", "center")
    ws.cell(row=gt_row2, column=start_col, value="Grand Total")
    for ci in range(start_col + 1, start_col + len(type_cols) + 1):
        col_letter = get_column_letter(ci)
        ws.cell(row=gt_row2, column=ci,
                value=f"=SUM({col_letter}3:{col_letter}{gt_row2-1})")


# ════════════════════════════════════════════════════════════════════════════
# FILE 2 — Annual FY Engineering Sheet
# Matches: Engineering_Timesheet_FY_2024-2025.xlsx
# ════════════════════════════════════════════════════════════════════════════

# Financial year months in order (Apr = start of FY)
FY_MONTHS = ["Apr", "May", "Jun", "Jul", "Aug", "Sep",
             "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]

def generate_fy_engineering_report(tickets, fy_label="2024-2025"):
    """
    Produces one sheet per POD, with:
    Row:    Resource Name | Emp Code | Product | Task | Client
    Cols:   Monthly hours for each month in the FY
    Matches Engineering_Timesheet_FY_2024-2025.xlsx exactly.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Group tickets by POD
    by_pod = defaultdict(list)
    for t in tickets:
        by_pod[t["pod"]].append(t)

    for pod, pod_tickets in sorted(by_pod.items()):
        _fy_pod_sheet(wb, pod, pod_tickets, fy_label)

    # Add a combined sheet — matches "DPAI-EDM-SNP-INFOSEC-DEVOPS" flat sheet
    _fy_combined_sheet(wb, tickets)

    os.makedirs("/tmp/reports", exist_ok=True)
    path = f"/tmp/reports/engineering_timesheet_FY_{fy_label}.xlsx"
    wb.save(path)
    return path


def _fy_pod_sheet(wb, pod_name, tickets, fy_label):
    ws = wb.create_sheet(pod_name[:31])  # Excel sheet name limit

    # Row 1: Version row (left blank — teams fill this in manually)
    ws.row_dimensions[1].height = 16
    v_cell = ws.cell(row=1, column=5, value=f"{pod_name} Version")
    v_cell.font = _font(bold=True, color="7F7F7F", size=9)

    # Row 2: FY label
    ws.row_dimensions[2].height = 16
    fy_cell = ws.cell(row=2, column=5, value=f"FY {fy_label}")
    fy_cell.font = _font(bold=True, color="7F7F7F", size=9)

    # Row 3: Column headers — matches exactly
    ws.row_dimensions[3].height = 22
    fixed_headers = ["Resource Name", "Emp Code", "Product", "Task", "Client"]
    fixed_widths  = [24, 12, 16, 18, 18]

    for ci, (h, w) in enumerate(zip(fixed_headers, fixed_widths), 1):
        _header_cell(ws, 3, ci, h, bg="1F4E79", fg="FFFFFF", bold=True,
                     size=10, h_align="center")
        _set_col_width(ws, get_column_letter(ci), w)

    for ci, month in enumerate(FY_MONTHS, start=6):
        _header_cell(ws, 3, ci, month, bg="2E75B6", fg="FFFFFF",
                     bold=True, size=10, h_align="center")
        _set_col_width(ws, get_column_letter(ci), 8)

    # Total column
    total_col = 6 + len(FY_MONTHS)
    _header_cell(ws, 3, total_col, "Total", bg="1F4E79", fg="FFFFFF",
                 bold=True, size=10, h_align="center")
    _set_col_width(ws, get_column_letter(total_col), 10)

    # Build data: group by (assignee, pod, issue_type, client)
    # aggregate hours per month
    # Key: (name, emp_code, product, task, client) → {month: hours}
    agg = defaultdict(lambda: defaultdict(float))

    for t in tickets:
        key = (
            t["assignee"],
            "",           # emp code — not in Jira, leave blank
            t["pod"],     # product = pod
            t["issue_type"] or "Development",
            t["client"],
        )
        if t["worklogs"]:
            for w in t["worklogs"]:
                month_abbr = _month_abbr(w["date"])
                if month_abbr:
                    agg[key][month_abbr] += w["hours"]
        else:
            month_abbr = _month_abbr(t["updated"])
            if month_abbr:
                agg[key][month_abbr] += t["hours_spent"]

    # Write data rows
    ROW_COLORS = ["EBF3FB", "FFFFFF"]
    for ri, (key, month_data) in enumerate(sorted(agg.items()), start=4):
        name, emp, product, task, client = key
        bg = ROW_COLORS[ri % 2]
        ws.row_dimensions[ri].height = 16

        _data_cell(ws, ri, 1, name,    bg=bg, bold=True, color="1F4E79")
        _data_cell(ws, ri, 2, emp,     bg=bg, color="7F7F7F")
        _data_cell(ws, ri, 3, product, bg=bg)
        _data_cell(ws, ri, 4, task,    bg=bg)
        _data_cell(ws, ri, 5, client,  bg=bg, bold=True, color="2E75B6")

        for ci, month in enumerate(FY_MONTHS, start=6):
            val = round(month_data.get(month, 0), 2) or None
            _data_cell(ws, ri, ci, val, bg=bg, h_align="center", num_fmt="0")

        # Total formula
        start_col_letter = get_column_letter(6)
        end_col_letter   = get_column_letter(6 + len(FY_MONTHS) - 1)
        total_cell = ws.cell(row=ri, column=total_col,
                             value=f"=SUM({start_col_letter}{ri}:{end_col_letter}{ri})")
        total_cell.font      = _font(bold=True, color="1F4E79", size=9)
        total_cell.fill      = _fill(bg)
        total_cell.alignment = _align("center", "center")
        total_cell.border    = _border_thin()
        total_cell.number_format = "0"

    # Grand total row at the bottom
    last_data_row = len(agg) + 3
    gt_row = last_data_row + 1
    ws.row_dimensions[gt_row].height = 20

    for ci in range(1, total_col + 1):
        c = ws.cell(row=gt_row, column=ci)
        c.fill   = _fill("1F4E79")
        c.border = _border_thin()
        c.font   = _font(bold=True, color="FFFFFF", size=10)
        c.alignment = _align("center", "center")

    ws.cell(row=gt_row, column=1, value="Grand Total")
    for ci in range(6, total_col + 1):
        col_letter = get_column_letter(ci)
        ws.cell(row=gt_row, column=ci,
                value=f"=SUM({col_letter}4:{col_letter}{gt_row-1})")


def _fy_combined_sheet(wb, tickets):
    """
    Flat combined sheet matching 'DPAI-EDM-SNP-INFOSEC-DEVOPS':
    EmployeeName | POD | Date | Module | Feature | Type | Client | Hours | JIRA# | Remark
    """
    ws = wb.create_sheet("All PODs Combined")
    ws.freeze_panes = "A2"

    headers = ["EmployeeName", "POD", "Date", "Module", "Feature",
               "Type (Bug/Feature/Meeting)", "Client",
               "Time spent in (Hours)", "JIRA#", "Remark"]
    widths  = [24, 12, 12, 14, 40, 20, 16, 22, 12, 30]

    ws.row_dimensions[1].height = 22
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        _header_cell(ws, 1, ci, h, bg="1F4E79", fg="FFFFFF",
                     bold=True, size=10, h_align="center")
        _set_col_width(ws, get_column_letter(ci), w)

    ROW_COLORS = ["EBF3FB", "FFFFFF"]
    ri = 2
    for t in tickets:
        entries = t["worklogs"] if t["worklogs"] else [{
            "author":  t["assignee"],
            "date":    t["updated"],
            "hours":   t["hours_spent"],
            "comment": "",
        }]
        for w in entries:
            bg = ROW_COLORS[ri % 2]
            ws.row_dimensions[ri].height = 16
            date_val = w.get("date", "")
            if isinstance(date_val, str) and date_val:
                try:
                    date_val = datetime.strptime(date_val, "%Y-%m-%d")
                except Exception:
                    pass
            vals = [
                w.get("author", t["assignee"]),
                t["pod"], date_val, t["pod"], t["summary"],
                t["issue_type"] or "Feature", t["client"],
                w.get("hours", 0), t["key"], w.get("comment", ""),
            ]
            for ci, val in enumerate(vals, 1):
                cell = _data_cell(ws, ri, ci, val, bg=bg)
                if ci == 3 and isinstance(val, datetime):
                    cell.number_format = "DD-MMM-YY"
                if ci == 8:
                    cell.alignment = _align("center", "center")
                    cell.font = _font(bold=True, color="1F4E79", size=9)
            ri += 1

    # Total
    tc = ws.cell(row=ri, column=7, value="Grand Total")
    tc.font = _font(bold=True, color="FFFFFF", size=10)
    tc.fill = _fill("1F4E79")
    tc.alignment = _align("right", "center")
    tc.border = _border_thin()
    tv = ws.cell(row=ri, column=8, value=f"=SUM(H2:H{ri-1})")
    tv.font = _font(bold=True, color="FFFFFF", size=10)
    tv.fill = _fill("1F4E79")
    tv.alignment = _align("center", "center")
    tv.border = _border_thin()
    tv.number_format = "0.00"


def _month_abbr(date_str):
    """Convert YYYY-MM-DD to 3-letter month abbr matching FY_MONTHS list."""
    if not date_str:
        return None
    try:
        if isinstance(date_str, datetime):
            dt = date_str
        else:
            dt = datetime.strptime(str(date_str)[:10], "%Y-%m-%d")
        return dt.strftime("%b")  # e.g. "Oct"
    except Exception:
        return None
